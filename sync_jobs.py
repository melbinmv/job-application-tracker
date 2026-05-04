import os
import re
import json
import base64
import pickle
import time
from datetime import datetime
from html import unescape as html_unescape
from email.utils import parsedate_to_datetime

import google.generativeai as genai
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE       = "job_applications.xlsx"
SHEET_NAME       = "Applications"
REVIEW_SHEET     = "Needs Review"
TOKEN_FILE       = "token.pickle"
CREDENTIALS_FILE = "credentials.json"
SCOPES           = ["https://www.googleapis.com/auth/gmail.readonly"]

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL   = "gemma-3-4b-it"
BATCH_SIZE     = 5
BATCH_DELAY    = 3

GMAIL_QUERY = (
    'to:me -from:me ('
    'subject:("application received" OR "thank you for applying" OR '
    '"thanks for applying" OR "we received your application" OR '
    '"application confirmation" OR "your application" OR '
    '"application for" OR "thank you for your interest" OR '
    '"application feedback" OR "update on your application" OR '
    '"unfortunately" OR "not been successful" OR "not move forward" OR '
    '"position has been filled" OR "regret to inform" OR '
    '"not be progressing" OR "not shortlisted" OR '
    '"decided not to proceed"))'
)


def get_gmail_service():
    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "wb") as f:
            pickle.dump(creds, f)
    return build("gmail", "v1", credentials=creds)


def fetch_messages(service, max_results: int = 200) -> list:
    response = service.users().messages().list(
        userId="me", q=GMAIL_QUERY, maxResults=max_results
    ).execute()
    return response.get("messages", [])


def _decode_part_data(data: str, encoding: str = "") -> str:
    if not data:
        return ""
    try:
        raw_bytes = base64.urlsafe_b64decode(data)
        if encoding.lower() == "quoted-printable":
            import quopri
            raw_bytes = quopri.decodestring(raw_bytes)
        return raw_bytes.decode("utf-8", errors="ignore")
    except Exception:
        return ""


def _strip_html(html: str) -> str:
    text = re.sub(r'<style[^>]*>.*?</style>', ' ', html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<script[^>]*>.*?</script>', ' ', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<!--.*?-->', ' ', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = html_unescape(text)
    text = text.replace('\xa0', ' ')
    text = re.sub(r'[\u200b\u200c\u200d\ufeff]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def get_email_body(msg_payload):
    mime    = msg_payload.get("mimeType", "")
    data    = msg_payload.get("body", {}).get("data", "")
    headers = {h["name"].lower(): h["value"] for h in msg_payload.get("headers", [])}
    enc     = headers.get("content-transfer-encoding", "")

    if mime == "text/plain" and data:
        return _decode_part_data(data, enc)
    if mime == "text/html" and data:
        return _strip_html(_decode_part_data(data, enc))

    plain, html = "", ""
    for part in msg_payload.get("parts", []):
        pm = part.get("mimeType", "")
        pd = part.get("body", {}).get("data", "")
        ph = {h["name"].lower(): h["value"] for h in part.get("headers", [])}
        pe = ph.get("content-transfer-encoding", "")
        if pm == "text/plain" and pd:
            plain = _decode_part_data(pd, pe)
        elif pm == "text/html" and pd:
            html = _strip_html(_decode_part_data(pd, pe))
        else:
            result = get_email_body(part)
            if result and not plain:
                plain = result
    return plain or html or ""


def fetch_email_data(service, msg_id: str) -> dict:
    msg     = service.users().messages().get(userId="me", id=msg_id, format="full").execute()
    headers = {h["name"].lower(): h["value"] for h in msg["payload"].get("headers", [])}
    try:
        date_obj = parsedate_to_datetime(
            re.sub(r'\s*\([^)]*\)', '', headers.get("date", "")).strip()
        ).date()
    except Exception:
        date_obj = datetime.today().date()

    return {
        "date":    date_obj,
        "sender":  headers.get("from", ""),
        "subject": headers.get("subject", ""),
        "body":    get_email_body(msg["payload"])[:1000],
    }


BATCH_PROMPT = """You are parsing job application emails. For each email below, extract:
- "company": actual employer name (not ATS platforms like Workable, Greenhouse, Indeed, Lever, Talos)
- "job_role": specific job title or "N/A" if not found
- "status": "Pending" (acknowledgment) or "Rejected" (rejection)

{emails}

Return ONLY a JSON array with exactly {count} objects in this format:
[
  {{"id": 1, "company": "Google", "job_role": "Data Analyst", "status": "Pending"}},
  {{"id": 2, "company": "Meta", "job_role": "N/A", "status": "Rejected"}}
]"""


def call_gemini_batch(email_batch: list) -> list:
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel(GEMINI_MODEL)

    emails_text = ""
    for i, email in enumerate(email_batch, 1):
        emails_text += (
            f"\n--- EMAIL {i} ---\n"
            f"SENDER: {email['sender']}\n"
            f"SUBJECT: {email['subject']}\n"
            f"BODY: {email['body'][:600]}\n"
        )

    prompt = BATCH_PROMPT.format(emails=emails_text, count=len(email_batch))

    for attempt in range(3):
        try:
            response = model.generate_content(
                prompt,
                generation_config=genai.GenerationConfig(temperature=0, max_output_tokens=512),
            )
            break
        except Exception as retry_err:
            wait_match = re.search(r'retry in (\d+)', str(retry_err), re.IGNORECASE)
            wait_secs  = int(wait_match.group(1)) + 2 if wait_match else 30
            print(f"   ⏳ Rate limited — waiting {wait_secs}s before retry {attempt+1}/3...")
            time.sleep(wait_secs)
    else:
        print(f"   ⚠️  Gemini failed after 3 retries")
        return [{"company": "Unknown", "job_role": "N/A", "status": "Pending"}] * len(email_batch)
    try:
        response = response
        text = response.text.strip()

        # Strip markdown code fences if present
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text).strip()

        # Extract JSON array if Gemini added extra text around it
        array_match = re.search(r'\[.*\]', text, re.DOTALL)
        if array_match:
            text = array_match.group()

        try:
            results = json.loads(text)
        except json.JSONDecodeError as je:
            print(f"   ⚠️  Gemini returned invalid JSON: {je}")
            print(f"   Raw response: {text[:300]}")
            return [{"company": "Unknown", "job_role": "N/A", "status": "Pending"}] * len(email_batch)

        status_map = {"rejected": "Rejected", "pending": "Pending"}
        parsed = []
        for r in results:
            parsed.append({
                "company":  str(r.get("company", "Unknown")).strip() or "Unknown",
                "job_role": str(r.get("job_role", "N/A")).strip() or "N/A",
                "status":   status_map.get(str(r.get("status", "")).lower(), "Pending"),
            })
        return parsed

    except Exception as e:
        print(f"   ⚠️  Gemini batch error: {e}")
        return [{"company": "Unknown", "job_role": "N/A", "status": "Pending"}] * len(email_batch)


HEADERS    = ["Date", "Company", "Job Role", "Status"]
COL_WIDTHS = [18, 25, 30, 15]

STATUS_STYLES = {
    "Pending":  {"color": "C47A1E"},
    "Rejected": {"color": "C0392B"},
}

STATUS_PRIORITY = {"Pending": 0, "Rejected": 1}


def _header_style(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", start_color="2E4057")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FFFFFF")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_headers(ws):
    ws.append(HEADERS)
    for col, cell in enumerate(ws[1], 1):
        _header_style(cell)
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS[col - 1]


def ensure_workbook() -> Workbook:
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _apply_headers(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _apply_headers(ws)
        ws.freeze_panes = "A2"

    if REVIEW_SHEET not in wb.sheetnames:
        rv = wb.create_sheet(REVIEW_SHEET)
        rv.append(["Date", "Sender", "Subject", "Body", "Extracted Company", "Extracted Job Role"])
        for col, cell in enumerate(rv[1], 1):
            _header_style(cell)
            rv.column_dimensions[cell.column_letter].width = [18, 30, 40, 80, 25, 25][col - 1]
        rv.freeze_panes = "A2"

    return wb


def append_review_row(ws, email: dict, entry: dict):
    ws.append([
        entry["date"],
        email.get("sender", ""),
        email.get("subject", ""),
        email.get("body", "")[:2000],
        entry["company"],
        entry["job_role"],
    ])
    r    = ws.max_row
    fill = PatternFill("solid", start_color="FFF3CD" if r % 2 == 0 else "FFFDE7")
    for col in range(1, 7):
        cell = ws.cell(row=r, column=col)
        cell.fill      = fill
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True if col == 4 else False)
    ws.row_dimensions[r].height = 60


def load_existing_rows(wb: Workbook) -> dict:
    ws = wb[SHEET_NAME]
    rows = {}
    for r in range(2, ws.max_row + 1):
        company = ws.cell(row=r, column=2).value
        if company:
            rows[str(company).strip().lower()] = r
    return rows


def update_status_in_row(ws, row_num: int, new_status: str):
    style = STATUS_STYLES.get(new_status, {"color": "2E4057"})
    cell  = ws.cell(row=row_num, column=4)
    cell.value     = new_status
    cell.font      = Font(name="Arial", size=10, bold=True, color=style["color"])
    cell.alignment = Alignment(horizontal="center", vertical="center")


def append_row(ws, entry: dict):
    ws.append([entry["date"], entry["company"], entry["job_role"], entry["status"]])
    r    = ws.max_row
    fill = PatternFill("solid", start_color="F0F4F8" if r % 2 == 0 else "FFFFFF")
    for col in range(1, 5):
        cell = ws.cell(row=r, column=col)
        cell.fill = fill
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
    style = STATUS_STYLES.get(entry["status"], {"color": "2E4057"})
    sc = ws.cell(row=r, column=4)
    sc.font      = Font(name="Arial", size=10, bold=True, color=style["color"])
    sc.alignment = Alignment(horizontal="center", vertical="center")


def main():
    if not GEMINI_API_KEY:
        print("❌ GEMINI_API_KEY not set. Run: export GEMINI_API_KEY='your_key_here'")
        return

    print("🔐 Authenticating with Gmail...")
    service = get_gmail_service()
    print("✅ Authenticated.\n")

    print("📬 Fetching job-related emails...")
    messages = fetch_messages(service)
    print(f"   Found {len(messages)} matching emails.\n")

    print("📥 Fetching email contents...")
    email_data = []
    for i, msg_meta in enumerate(messages):
        try:
            email_data.append(fetch_email_data(service, msg_meta["id"]))
        except Exception as e:
            print(f"   ⚠️  Error fetching email {i+1}: {e}")
            email_data.append(None)

    wb            = ensure_workbook()
    ws            = wb[SHEET_NAME]
    existing_rows = load_existing_rows(wb)

    added    = 0
    updated  = 0
    skipped  = 0
    reviewed = 0

    total_batches = (len(email_data) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"\n🤖 Sending to Gemini in {total_batches} batches of {BATCH_SIZE}...\n")

    for batch_num in range(total_batches):
        start = batch_num * BATCH_SIZE
        end   = start + BATCH_SIZE
        batch_emails  = [e for e in email_data[start:end] if e is not None]
        batch_indices = [i for i, e in enumerate(email_data[start:end], start) if e is not None]

        if not batch_emails:
            continue

        print(f"   Batch {batch_num + 1}/{total_batches} ({len(batch_emails)} emails)...")
        results = call_gemini_batch(batch_emails)

        rv = wb[REVIEW_SHEET]
        for email, result in zip(batch_emails, results):
            entry = {
                "date":     email["date"],
                "company":  result["company"],
                "job_role": result["job_role"],
                "status":   result["status"],
            }

            company  = entry["company"].strip()
            job_role = entry["job_role"].strip()
            needs_review = (
                company.lower() in ("unknown", "", "n/a") or
                job_role.lower() in ("n/a", "", "unknown")
            )

            if needs_review:
                append_review_row(rv, email, entry)
                reviewed += 1
                print(f"      📋 Needs review: {company} — {job_role}")
                continue

            company_key = company.lower()
            status_icon = "❌" if entry["status"] == "Rejected" else "✅"
            print(f"      {status_icon} {company} — {job_role} [{entry['status']}]")

            if company_key in existing_rows:
                row_num    = existing_rows[company_key]
                cur_status = ws.cell(row=row_num, column=4).value or "Pending"
                if STATUS_PRIORITY.get(entry["status"], 0) > STATUS_PRIORITY.get(cur_status, 0):
                    update_status_in_row(ws, row_num, entry["status"])
                    updated += 1
                else:
                    skipped += 1
            else:
                append_row(ws, entry)
                existing_rows[company_key] = ws.max_row
                added += 1

        if batch_num < total_batches - 1:
            print(f"   ⏳ Waiting {BATCH_DELAY}s before next batch...")
            time.sleep(BATCH_DELAY)

    wb.save(EXCEL_FILE)
    print(f"\n{'='*55}")
    print(f"   ✅ Added:          {added}")
    print(f"   🔄 Updated:        {updated}")
    print(f"   ⏭  Skipped:        {skipped}")
    print(f"   📋 Needs review:   {reviewed} (see '{REVIEW_SHEET}' sheet)")
    print(f"   🤖 Gemini calls:   {total_batches} (batched from {len(messages)} emails)")
    print(f"   📁 Saved to: {os.path.abspath(EXCEL_FILE)}")


if __name__ == "__main__":
    main()